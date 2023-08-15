import os
from openpyxl import load_workbook
from .conftest import RESOURCE_ROOT_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_cell_value():
    xlsx_file_path = os.path.join(RESOURCE_ROOT_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_file_path)
    sheet = workbook.active

    row = 3
    col = 2

    expected_value = 'Mara'

    assert sheet.cell(row=row, column=col).value == expected_value
